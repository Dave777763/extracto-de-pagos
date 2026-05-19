"""
Comprobantes de Pagos de Impuestos SAT
Una fila por PDF · Impuestos en columnas fijas
"""
import os, re, io, uuid
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import fitz
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "sat-pagos-2024"
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ── Catálogo ordenado de impuestos ──────────────────────────────────────────
# Portal Nuevo (4 conceptos)
PORTAL_NUEVO_IMPUESTOS = [
    "ISR personas morales",
    "ISR retenciones por salarios",
    "Impuesto al Valor Agregado. Personas morales. Región Fronteriza",
    "IVA retenciones",
]
# Portal Anterior (3 conceptos)
PORTAL_ANTERIOR_IMPUESTOS = [
    "ISR RETENCIONES POR SERVICIOS PROFESIONALES/RÉGIMEN SIMPLIFICADO DE CONFIANZA",
    "ISR POR PAGOS POR CUENTA DE TERCEROS O RETENCIONES POR ARRENDAMIENTO DE INMUEBLES",
    "ISR RETENCIONES POR INTERESES",
]

ALL_IMPUESTOS = PORTAL_NUEVO_IMPUESTOS + PORTAL_ANTERIOR_IMPUESTOS

# Alias → nombre canónico
ALIAS = {
    "isr personas morales":                         "ISR personas morales",
    "sr personas morales":                          "ISR personas morales",
    "isr retenciones por salarios":                 "ISR retenciones por salarios",
    "impuesto al valor agregado. personas morales. región fronteriza": "Impuesto al Valor Agregado. Personas morales. Región Fronteriza",
    "impuesto al valor agregado. personas morales. region fronteriza":  "Impuesto al Valor Agregado. Personas morales. Región Fronteriza",
    "iva retenciones":                              "IVA retenciones",
    "isr retenciones por servicios profesionales/régimen simplificado de confianza": "ISR RETENCIONES POR SERVICIOS PROFESIONALES/RÉGIMEN SIMPLIFICADO DE CONFIANZA",
    "isr retenciones por servicios profesionales":  "ISR RETENCIONES POR SERVICIOS PROFESIONALES/RÉGIMEN SIMPLIFICADO DE CONFIANZA",
    "isr por pagos por cuenta de terceros o retenciones por arrendamiento de inmuebles": "ISR POR PAGOS POR CUENTA DE TERCEROS O RETENCIONES POR ARRENDAMIENTO DE INMUEBLES",
    "isr por pagos por cuenta de terceros":         "ISR POR PAGOS POR CUENTA DE TERCEROS O RETENCIONES POR ARRENDAMIENTO DE INMUEBLES",
    "retenciones por arrendamiento de inmuebles":   "ISR POR PAGOS POR CUENTA DE TERCEROS O RETENCIONES POR ARRENDAMIENTO DE INMUEBLES",
    "isr retenciones por intereses":                "ISR RETENCIONES POR INTERESES",
}

# Sub-columnas por cada impuesto
SUB_COLS = ["A cargo", "Parte actualizada", "Recargos",
            "Fecha pago anterior", "Monto pago anterior",
            "Cantidad a cargo", "Cantidad a favor", "Compensaciones", "Cantidad a pagar"]
SUB_KEYS = ["a_cargo", "parte_actualizada", "recargos",
            "fecha_pago_anterior", "monto_pago_anterior",
            "cantidad_a_cargo", "cantidad_a_favor", "compensaciones", "cantidad_a_pagar"]
SUB_NUMERIC = {0, 1, 2, 4, 5, 6, 7, 8}   # índices numéricos en SUB_KEYS

# ── Configuración de Columnas ──
# (Coincide con index.html)
HDR_KEYS = ["archivo", "ruta_archivo", "periodo", "ejercicio", "fecha_presentacion", "tipo_declaracion", "tipo_complementaria", "num_operacion_declaracion", "vencimiento_obligacion"]
HDR_COLS = ["Nombre del archivo", "Ruta del archivo", "Período", "Ejercicio", "Fecha presentación", "Tipo decl.", "Tipo complem.", "No. op. declaración", "Venc. obligación"]

# Columnas de cierre (después de impuestos)
TAIL_COLS = ["Importe total a pagar", "Línea de captura", "Vigente hasta",
             "Institución de crédito", "Fecha del pago",
             "No. de Operación pago", "Llave de pago"]
TAIL_KEYS = ["importe_total", "linea_captura", "vigente_hasta",
             "institucion_credito", "fecha_pago",
             "no_operacion_pago", "llave_pago"]
TAIL_NUMERIC = {0}

# ── Helpers ─────────────────────────────────────────────────────────────────
def parse_amount(t):
    if not t: return None
    t = str(t).replace("$","").replace(",","").strip()
    # Formato contable con paréntesis: (1079) → -1079
    if t.startswith('(') and t.endswith(')'):
        try: return -float(t[1:-1])
        except: return None
    try: return float(t)
    except: return None

def ef(text, label):
    m = re.search(rf'{label}\s*:?\s*\n?\s*(.+)', text, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def normalize(raw: str) -> str:
    """Normaliza texto: quita espacios múltiples y ajusta tildes según sea necesario."""
    if not raw: return ""
    raw = re.sub(r'\s+', ' ', raw)
    # Correcciones de OCR / caracteres raros reportados
    raw = re.sub(r'Regi.n\s+Fronteriza', 'Región Fronteriza', raw, flags=re.IGNORECASE)
    
    k = raw.lower().strip()
    if k in ALIAS: return ALIAS[k]
    for a, v in ALIAS.items():
        if a in k or k in a: return v
    return raw.strip()

# ── Parse PDF → 1 dict por PDF ──────────────────────────────────────────────
def parse_pdf(pdf_bytes: bytes, filename: str, filepath: str = "") -> dict:
    """Extrae datos de un PDF usando PyMuPDF (fitz) + regex."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = "".join(p.get_text() + "\n" for p in doc)
    doc.close()

    rec = {
        "archivo": filename,
        "ruta_archivo": filepath,
        "rfc":        ef(text, "RFC"),
        "razon_social": ef(text, r"Denominaci[oó]n o raz[oó]n social"),
        "periodo":    ef(text, r"Per[ií]odo de la declaraci[oó]n"),
        "ejercicio":  ef(text, "Ejercicio"),
        "tipo_declaracion": ef(text, r"Tipo de declaraci[oó]n"),
        "tipo_complementaria": ef(text, r"Tipo de Complementaria"),
        "num_operacion_declaracion": ef(text, r"N[uú]mero de operaci[oó]n"),
    }

    m = re.search(r'Fecha y hora de presentaci[oó]n\s*:\s*\n?\s*(\S+\s+\S+)', text, re.IGNORECASE)
    rec["fecha_presentacion"] = m.group(1).strip() if m else ""

    # Pago bancario — busca en la sección "INFORMACIÓN DEL PAGO RECIBIDO"
    ms = re.search(r'INFORMACI[OÓ]N DEL PAGO RECIBIDO.+', text, re.IGNORECASE | re.DOTALL)
    if ms:
        pago_section = ms.group(0)[:600]
        
        # Institución de crédito: puede aparecer dos veces; la 2a tiene el nombre real del banco
        banco_matches = re.findall(r'Instituci[oó]n de\s*\n?cr[eé]dito\s*:?\s*\n?\s*([^\n\r]+)', pago_section, re.IGNORECASE)
        banco = ""
        for bm in banco_matches:
            if "AUTORIZADA" not in bm.upper():
                banco = bm.strip()
                break
        rec["institucion_credito"] = banco

        m_fp = re.search(r'Fecha del\s*\n?pago\s*:?\s*\n?\s*([\d/]+)', pago_section, re.IGNORECASE)
        rec["fecha_pago"] = m_fp.group(1).strip() if m_fp else ""

        m_nop = re.search(r'No\.\s*de\s*\n?Operaci[oó]n\s*:?\s*\n?\s*(\S+)', pago_section, re.IGNORECASE)
        rec["no_operacion_pago"] = m_nop.group(1).strip() if m_nop else ""

        m_lp = re.search(r'Llave de\s*\n?Pago\s*:?\s*\n?\s*(\S+)', pago_section, re.IGNORECASE)
        rec["llave_pago"] = m_lp.group(1).strip() if m_lp else ""

        m_ip = re.search(r'Importe\s*\n?pagado\s*:?\s*\n?\s*\$?([\d,]+)', pago_section, re.IGNORECASE)
        rec["importe_total"] = parse_amount(m_ip.group(1)) if m_ip else None
    else:
        rec["institucion_credito"] = ""
        rec["fecha_pago"] = ""
        rec["no_operacion_pago"] = ""
        rec["llave_pago"] = ""
        rec["importe_total"] = None

    # Línea de captura: usar la del pago bancario (pago_section) si existe, más confiable
    m_lc = None
    if ms:
        m_lc = re.search(r'L[ií]nea de\s*\n?Captura\s*:\s*\n?\s*([^\n\r]+)', pago_section, re.IGNORECASE)
    if not m_lc:  # fallback: buscar en texto completo
        m_lc = re.search(r'L[ií]nea de\s*\n?Captura\s*:\s*\n?\s*([^\n\r]+)', text, re.IGNORECASE)
    rec["linea_captura"] = m_lc.group(1).strip() if m_lc else ""

    m_vig = re.search(r'Vigente hasta\s*:?\s*\n?\s*([\d/]+)', text, re.IGNORECASE)
    rec["vigente_hasta"] = m_vig.group(1).strip() if m_vig else ""

    m_venc = re.search(r'Vencimiento Obligaci[oó]n\s*:?\s*\n?\s*([\d/]+)', text, re.IGNORECASE)
    rec["vencimiento_obligacion"] = m_venc.group(1).strip() if m_venc else ""

    # Inicializa todos los impuestos en None
    for imp in ALL_IMPUESTOS:
        for sk in SUB_KEYS:
            rec[f"{imp}|{sk}"] = None

    # Patrón para un bloque completo por concepto
    block_pattern = re.compile(
        r'(Concepto de pago\s+\d+\s*:\s*\n.+?)(?=Concepto de pago\s+\d+\s*:|INFORMACI[OÓ]N REGISTRADA|$)',
        re.DOTALL
    )
    raw_blocks = block_pattern.findall(text)

    for block_text in raw_blocks:
        # Nombre del concepto
        m_nombre = re.search(r'Concepto de pago\s+\d+\s*:\s*\n(.+?)\n', block_text)
        if not m_nombre:
            continue
        nombre = normalize(m_nombre.group(1))
        if nombre not in ALL_IMPUESTOS:
            continue

        def _get(pattern, txt=block_text):
            m = re.search(pattern, txt, re.IGNORECASE)
            return m.group(1).strip() if m else None

        a_cargo          = parse_amount(_get(r'A cargo:\s*\n([\d,]+)'))
        a_favor_raw      = _get(r'A favor:\s*\n([\d,]+)')
        parte_act        = parse_amount(_get(r'Parte actualizada:\s*\n([\d,]+)'))
        recargos         = parse_amount(_get(r'Recargos:\s*\n([\d,]+)'))
        fecha_pago_ant   = _get(r'Fecha del pago realizado con anterioridad:\s*\n(.+?)')
        monto_pago_ant   = parse_amount(_get(r'Monto pagado con anterioridad:\s*\n([\d,]+)'))
        cant_cargo       = parse_amount(_get(r'Cantidad a cargo:\s*\n([\d,]+)'))
        cant_favor_raw   = _get(r'Cantidad a favor:\s*\n([\d,]+)')
        cant_favor       = parse_amount(cant_favor_raw) or parse_amount(a_favor_raw)
        # Compensaciones: puede ser (1,079) o 1,079
        comp_raw         = _get(r'Compensaci[oó]n(?:es)?:\s*\n(\(?[\d,]+\)?)')
        compensaciones   = parse_amount(comp_raw)
        cant_pagar       = parse_amount(_get(r'Cantidad a pagar:\s*\n([\d,]+)'))

        vals = [a_cargo, parte_act, recargos, fecha_pago_ant, monto_pago_ant,
                cant_cargo, cant_favor, compensaciones, cant_pagar]
        for sk, v in zip(SUB_KEYS, vals):
            rec[f"{nombre}|{sk}"] = v

    # Portal detectado
    portales = set()
    for imp in PORTAL_NUEVO_IMPUESTOS:
        if rec.get(f"{imp}|a_cargo") is not None:
            portales.add("Portal Nuevo")
    for imp in PORTAL_ANTERIOR_IMPUESTOS:
        if rec.get(f"{imp}|a_cargo") is not None:
            portales.add("Portal Anterior")
    rec["portal"] = " + ".join(portales) if portales else "Desconocido"

    return rec

# ── Colores ──────────────────────────────────────────────────────────────────
C_NUEVO    = "1a6b3a"
C_ANTERIOR = "1a3a6b"
C_WHITE    = "FFFFFF"
C_NUEVO_L  = "C8E6C9"
C_ANT_L    = "BBDEFB"
C_HDR_BG   = "37474F"
C_TOTAL    = "F57F17"

def brd():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(cell, bg=None, bold=False, color="000000",
               halign="left", valign="center", wrap=False, num_fmt=None, size=9):
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = brd()
    if num_fmt:
        cell.number_format = num_fmt

def tiene_impuesto(rec, imp):
    for sk in SUB_KEYS:
        val = rec.get(f"{imp}|{sk}")
        if val is not None and val != "":
            return True
    return False

def get_date_key(rec):
    fp = rec.get("fecha_presentacion", "")
    if not fp:
        return datetime.min
    try:
        return datetime.strptime(fp, "%d/%m/%Y %H:%M")
    except:
        try:
            return datetime.strptime(fp, "%d/%m/%Y")
        except:
            return datetime.min

MESES_MAP = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
}

def get_period_sort_key(rec):
    try:
        ej = int(rec.get("ejercicio", 0))
    except:
        ej = 0
    pe = str(rec.get("periodo", "")).lower().strip()
    month_num = MESES_MAP.get(pe, 99)
    try:
        dt = get_date_key(rec)
    except:
        dt = datetime.min
    return (ej, month_num, dt)

def render_tax_table(ws, start_row, title, records, imp):
    ws.cell(start_row, 11, value=title)
    bg_hdr = C_NUEVO if imp in PORTAL_NUEVO_IMPUESTOS else C_ANTERIOR
    cell_style(ws.cell(start_row, 11), bg_hdr, True, C_WHITE, "center", size=8)
    ws.merge_cells(start_row=start_row, start_column=11, end_row=start_row, end_column=18)
    
    meta_labels = ["Nombre del archivo", "Ruta del archivo", "Período", "RFC", "Ejercicio", "Fecha presentación", "Tipo decl.", "Tipo complem.", "No. op. declaración", "Venc. obligación"]
    for ci, label in enumerate(meta_labels, 1):
        ws.cell(start_row, ci, value=label)
        cell_style(ws.cell(start_row, ci), C_HDR_BG, True, C_WHITE, "center", wrap=True, size=8)
        ws.merge_cells(start_row=start_row, start_column=ci, end_row=start_row+1, end_column=ci)
        cell_style(ws.cell(start_row+1, ci), C_HDR_BG, False, C_WHITE)
        
    for ci, label in enumerate(SUB_COLS, 11):
        ws.cell(start_row+1, ci, value=label)
        cell_style(ws.cell(start_row+1, ci), bg_hdr, True, C_WHITE, "center", wrap=True, size=7)
        
    for ci, label in enumerate(TAIL_COLS, 19):
        ws.cell(start_row, ci, value=label)
        cell_style(ws.cell(start_row, ci), C_HDR_BG, True, C_WHITE, "center", wrap=True, size=8)
        ws.merge_cells(start_row=start_row, start_column=ci, end_row=start_row+1, end_column=ci)
        cell_style(ws.cell(start_row+1, ci), C_HDR_BG, False, C_WHITE)
        
    ws.row_dimensions[start_row].height = 20
    ws.row_dimensions[start_row+1].height = 25
    
    current_row = start_row + 2
    for ri, rec in enumerate(records, start=current_row):
        portal = rec.get("portal", "")
        row_bg = "F1F8E9" if "Nuevo" in portal and "Anterior" not in portal else \
                 "E3F2FD" if "Anterior" in portal and "Nuevo" not in portal else "FFF8E1"
                 
        meta_keys = ["archivo", "ruta_archivo", "periodo", "rfc", "ejercicio", "fecha_presentacion", "tipo_declaracion", "tipo_complementaria", "num_operacion_declaracion", "vencimiento_obligacion"]
        for ci, key in enumerate(meta_keys, 1):
            cell = ws.cell(ri, ci, value=rec.get(key, ""))
            cell_style(cell, row_bg, halign="center")
            
        for ci, sk in enumerate(SUB_KEYS, 11):
            val = rec.get(f"{imp}|{sk}")
            cell = ws.cell(ri, ci, value=val)
            if (ci - 11) in SUB_NUMERIC and val is not None:
                cell_style(cell, row_bg, halign="right", num_fmt='#,##0.00')
            else:
                cell_style(cell, row_bg, halign="center")
                
        for ci, tk in enumerate(TAIL_KEYS, 19):
            val = rec.get(tk)
            cell = ws.cell(ri, ci, value=val)
            if (ci - 19) in TAIL_NUMERIC and val is not None:
                cell_style(cell, row_bg, bold=True, halign="right", num_fmt='#,##0.00')
            else:
                cell_style(cell, row_bg, halign="center")
        ws.row_dimensions[ri].height = 16
        current_row += 1
        
    tr = current_row
    ws.cell(tr, 1, value="TOTALES")
    cell_style(ws.cell(tr, 1), C_TOTAL, True, C_WHITE, "center")
    for ci in range(2, 11):
        cell_style(ws.cell(tr, ci), C_TOTAL, True, C_WHITE)
        
    for ci in range(11, 19):
        cell = ws.cell(tr, ci)
        if (ci - 11) in SUB_NUMERIC and len(records) > 0:
            col_letter = get_column_letter(ci)
            cell.value = f"=SUM({col_letter}{start_row+2}:{col_letter}{tr-1})"
            cell_style(cell, C_TOTAL, True, C_WHITE, "right", num_fmt='#,##0.00')
        else:
            cell_style(cell, C_TOTAL, True, C_WHITE)
            
    cell = ws.cell(tr, 19)
    if len(records) > 0:
        col_letter = get_column_letter(19)
        cell.value = f"=SUM({col_letter}{start_row+2}:{col_letter}{tr-1})"
        cell_style(cell, C_TOTAL, True, C_WHITE, "right", num_fmt='#,##0.00')
    else:
        cell_style(cell, C_TOTAL, True, C_WHITE)
        
    for ci in range(20, 26):
        cell_style(ws.cell(tr, ci), C_TOTAL, True, C_WHITE)
        
    ws.row_dimensions[tr].height = 18
    return tr

# ── Helpers para identificar pagos pendientes / atrasados ───────────────────
def get_payment_year(rec):
    fp = rec.get("fecha_pago", "")
    if not fp:
        return None
    # El formato suele ser "DD/MM/YYYY" o "DD/MM/YYYY HH:MM"
    m = re.search(r'\d{2}/\d{2}/(\d{4})', fp)
    if m:
        try:
            return int(m.group(1))
        except:
            pass
    for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(fp, fmt).year
        except:
            continue
    return None

def get_exercise_year(rec):
    ej = rec.get("ejercicio", "")
    if not ej:
        return None
    try:
        return int(ej)
    except:
        return None

def is_pending_payment(rec):
    # 1) Impuestos que tengan "Cantidad a pagar" sin datos bancarios (fecha_pago)
    if not tiene_banco(rec):
        for imp in ALL_IMPUESTOS:
            val = rec.get(f"{imp}|cantidad_a_pagar")
            if val is not None and isinstance(val, (int, float)) and val > 0:
                return True

    # 2) Impuestos con fecha de pago el siguiente ejercicio (año de pago > año de ejercicio)
    pay_yr = get_payment_year(rec)
    ex_yr = get_exercise_year(rec)
    if pay_yr and ex_yr and pay_yr > ex_yr:
        return True

    return False

# ── Renderizar hoja completa de registros ────────────────────────────────────
def write_full_record_sheet(ws, records):
    ws.freeze_panes = "A4"

    from collections import defaultdict
    groups = defaultdict(list)
    for r in records:
        ej = str(r.get("ejercicio", "")).strip() or "SinEjercicio"
        rfc = str(r.get("rfc", "")).strip() or "SinRFC"
        groups[(ej, rfc)].append(r)

    sorted_groups = sorted(groups.items(), key=lambda x: (x[0][1], x[0][0]))

    n_imp = len(ALL_IMPUESTOS)
    total_cols = len(HDR_COLS) + n_imp * len(SUB_KEYS) + len(TAIL_COLS)

    # ── Anchos de columna ──
    ws.column_dimensions[get_column_letter(1)].width = 12  # Período
    ws.column_dimensions[get_column_letter(2)].width = 8   # Ejercicio
    ws.column_dimensions[get_column_letter(3)].width = 18  # Fecha pres
    ws.column_dimensions[get_column_letter(4)].width = 14  # Tipo decl
    ws.column_dimensions[get_column_letter(5)].width = 20  # No. op
    base = len(HDR_COLS) + 1
    for i in range(n_imp * len(SUB_KEYS)):
        si = i % len(SUB_KEYS)
        w = 14 if si == 3 else 11  # fecha más ancha
        ws.column_dimensions[get_column_letter(base + i)].width = w
    for ti in range(len(TAIL_COLS)):
        ws.column_dimensions[get_column_letter(base + n_imp*len(SUB_KEYS) + ti)].width = 16

    current_row = 1
    for (ej, rfc), group_recs in sorted_groups:
        # ── Fila de título: RFC + Razón Social + Ejercicio ──
        rfcs = list(dict.fromkeys(r["rfc"] for r in group_recs if r.get("rfc")))
        rns  = list(dict.fromkeys(r["razon_social"] for r in group_recs if r.get("razon_social")))
        
        ws.cell(current_row, 1).value = "RFC:"
        cell_style(ws.cell(current_row, 1), C_HDR_BG, True, C_WHITE, size=10)
        ws.cell(current_row, 2).value = ", ".join(rfcs) if rfcs else "Sin RFC"
        cell_style(ws.cell(current_row, 2), C_HDR_BG, False, C_WHITE, size=10)
        ws.cell(current_row, 3).value = "RAZÓN SOCIAL:"
        cell_style(ws.cell(current_row, 3), C_HDR_BG, True, C_WHITE, size=10)
        ws.cell(current_row, 4).value = ", ".join(rns)
        cell_style(ws.cell(current_row, 4), C_HDR_BG, False, C_WHITE, size=10)
        ws.cell(current_row, 5).value = "EJERCICIO:"
        cell_style(ws.cell(current_row, 5), C_HDR_BG, True, C_WHITE, size=10)
        ws.cell(current_row, 6).value = ej
        cell_style(ws.cell(current_row, 6), C_HDR_BG, False, C_WHITE, size=10)

        hdr_row1 = current_row + 1
        hdr_row2 = current_row + 2

        # HDR (columnas iniciales)
        for ci, h in enumerate(HDR_COLS, 1):
            ws.cell(hdr_row1, ci, value=h)
            cell_style(ws.cell(hdr_row1, ci), C_HDR_BG, True, C_WHITE, "center", wrap=True, size=8)
            ws.merge_cells(start_row=hdr_row1, start_column=ci, end_row=hdr_row2, end_column=ci)
            cell_style(ws.cell(hdr_row2, ci), C_HDR_BG, False, C_WHITE)
        ws.row_dimensions[hdr_row1].height = 40

        # TAIL (columnas finales)
        tail_start_col = len(HDR_COLS) + n_imp * len(SUB_KEYS) + 1
        for ti, tc in enumerate(TAIL_COLS):
            tc_col = tail_start_col + ti
            ws.cell(hdr_row1, tc_col, value=tc)
            cell_style(ws.cell(hdr_row1, tc_col), C_HDR_BG, True, C_WHITE, "center", wrap=True, size=8)
            ws.merge_cells(start_row=hdr_row1, start_column=tc_col, end_row=hdr_row2, end_column=tc_col)
            cell_style(ws.cell(hdr_row2, tc_col), C_HDR_BG, False, C_WHITE)

        # Grupos de impuesto (hdr_row1)
        col = len(HDR_COLS) + 1
        for imp in ALL_IMPUESTOS:
            bg = C_NUEVO if imp in PORTAL_NUEVO_IMPUESTOS else C_ANTERIOR
            ws.cell(hdr_row1, col, value=imp)
            cell_style(ws.cell(hdr_row1, col), bg, True, C_WHITE, "center", wrap=True, size=8)
            end_col = col + len(SUB_KEYS) - 1
            ws.merge_cells(start_row=hdr_row1, start_column=col, end_row=hdr_row1, end_column=end_col)
            col = end_col + 1

        # Sub-columnas de impuesto (hdr_row2)
        col = len(HDR_COLS) + 1
        for imp in ALL_IMPUESTOS:
            bg_hdr = C_NUEVO if imp in PORTAL_NUEVO_IMPUESTOS else C_ANTERIOR
            for sc in SUB_COLS:
                ws.cell(hdr_row2, col, value=sc)
                cell_style(ws.cell(hdr_row2, col), bg_hdr, True, C_WHITE, "center", wrap=True, size=7)
                col += 1

        ws.row_dimensions[hdr_row2].height = 35

        # ── Datos ──
        data_start_row = hdr_row2 + 1
        for ri, rec in enumerate(group_recs, start=data_start_row):
            portal = rec.get("portal","")
            row_bg = "F1F8E9" if "Nuevo" in portal and "Anterior" not in portal else \
                     "E3F2FD" if "Anterior" in portal and "Nuevo" not in portal else "FFF8E1"

            col = 1
            for hk in HDR_KEYS:
                v = rec.get(hk, "")
                cell = ws.cell(ri, col, value=v)
                cell_style(cell, row_bg, halign="center")
                col += 1

            for imp in ALL_IMPUESTOS:
                for idx, sk in enumerate(SUB_KEYS):
                    v = rec.get(f"{imp}|{sk}")
                    cell = ws.cell(ri, col, value=v)
                    bg = row_bg
                    if idx in SUB_NUMERIC and v is not None:
                        cell_style(cell, bg, halign="right", num_fmt='#,##0.00')
                    else:
                        cell_style(cell, bg, halign="center")
                    col += 1

            for tidx, tk in enumerate(TAIL_KEYS):
                v = rec.get(tk)
                cell = ws.cell(ri, col, value=v)
                if tidx in TAIL_NUMERIC and v is not None:
                    cell_style(cell, row_bg, bold=True, halign="right", num_fmt='#,##0.00')
                else:
                    cell_style(cell, row_bg, halign="center")
                col += 1

            ws.row_dimensions[ri].height = 16

        # ── Fila TOTAL ──
        tr = data_start_row + len(group_recs)
        ws.cell(tr, 1, value="TOTALES")
        cell_style(ws.cell(tr,1), C_TOTAL, True, C_WHITE, "center")
        for c in range(2, len(HDR_COLS)+1):
            cell_style(ws.cell(tr,c), C_TOTAL, True, C_WHITE)

        col = len(HDR_COLS) + 1
        for imp in ALL_IMPUESTOS:
            for idx, sk in enumerate(SUB_KEYS):
                cell = ws.cell(tr, col)
                if idx in SUB_NUMERIC:
                    if len(group_recs) > 0:
                        col_letter = get_column_letter(col)
                        formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{tr-1})"
                        cell.value = formula
                    else:
                        cell.value = 0
                    cell_style(cell, C_TOTAL, True, C_WHITE, "right", num_fmt='#,##0.00')
                else:
                    cell_style(cell, C_TOTAL, True, C_WHITE)
                col += 1

        cell = ws.cell(tr, col)
        if len(group_recs) > 0:
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{tr-1})"
        else:
            cell.value = 0
        cell_style(cell, C_TOTAL, True, C_WHITE, "right", num_fmt='#,##0.00')
        for c in range(col+1, col+len(TAIL_COLS)):
            cell_style(ws.cell(tr, c), C_TOTAL, True, C_WHITE)

        ws.row_dimensions[tr].height = 18
        
        current_row = tr + 3

# ── Generar Excel ────────────────────────────────────────────────────────────
def generate_excel(records):
    wb = openpyxl.Workbook()
    
    # Pre-calcular RFCs y Razones Sociales para las hojas individuales de impuestos
    rfcs = list(dict.fromkeys(r["rfc"] for r in records if r.get("rfc")))
    rns  = list(dict.fromkeys(r["razon_social"] for r in records if r.get("razon_social")))

    # ── 1. HOJA PRINCIPAL: Pagos de Impuestos ──
    ws = wb.active
    ws.title = "Pagos de Impuestos"
    write_full_record_sheet(ws, records)

    # ── 2. HOJA ADICIONAL: X PAGAR ──
    x_pagar_records = [r for r in records if is_pending_payment(r)]
    ws_x_pagar = wb.create_sheet(title="X PAGAR")
    write_full_record_sheet(ws_x_pagar, x_pagar_records)

    # ── 3. HOJAS ADICIONALES POR IMPUESTO ──
    for imp in ALL_IMPUESTOS:
        imp_records = [r for r in records if tiene_impuesto(r, imp)]
        
        sheet_title = imp[:30]
        for char in "[]*:?/\\":
            sheet_title = sheet_title.replace(char, "")
        ws_imp = wb.create_sheet(title=sheet_title)
        
        ws_imp.column_dimensions[get_column_letter(1)].width = 25  # Nombre del archivo
        ws_imp.column_dimensions[get_column_letter(2)].width = 20  # Ruta del archivo
        ws_imp.column_dimensions[get_column_letter(3)].width = 12  # Período
        ws_imp.column_dimensions[get_column_letter(4)].width = 16  # RFC
        ws_imp.column_dimensions[get_column_letter(5)].width = 8   # Ejercicio
        ws_imp.column_dimensions[get_column_letter(6)].width = 18  # Fecha presentación
        ws_imp.column_dimensions[get_column_letter(7)].width = 14  # Tipo decl.
        ws_imp.column_dimensions[get_column_letter(8)].width = 20  # Tipo complem.
        ws_imp.column_dimensions[get_column_letter(9)].width = 18  # No. op. declaración
        ws_imp.column_dimensions[get_column_letter(10)].width = 14 # Venc. obligación
        for ci in range(11, 19):
            ws_imp.column_dimensions[get_column_letter(ci)].width = 13
        for ci in range(19, 26):
            ws_imp.column_dimensions[get_column_letter(ci)].width = 16

        from collections import defaultdict
        group_by_rfc_ej = defaultdict(list)
        for r in imp_records:
            ej = str(r.get("ejercicio", "")).strip() or "SinEjercicio"
            rfc = str(r.get("rfc", "")).strip() or "SinRFC"
            group_by_rfc_ej[(ej, rfc)].append(r)

        current_row_imp = 1
        for (ej, rfc), rfc_ej_records in sorted(group_by_rfc_ej.items(), key=lambda x: (x[0][1], x[0][0])):
            rfcs = list(dict.fromkeys(r["rfc"] for r in rfc_ej_records if r.get("rfc")))
            rns  = list(dict.fromkeys(r["razon_social"] for r in rfc_ej_records if r.get("razon_social")))
            ws_imp.cell(current_row_imp,1).value = "RFC:"
            cell_style(ws_imp.cell(current_row_imp,1), C_HDR_BG, True, C_WHITE, size=10)
            ws_imp.cell(current_row_imp,2).value = ", ".join(rfcs) if rfcs else "Sin RFC"
            cell_style(ws_imp.cell(current_row_imp,2), C_HDR_BG, False, C_WHITE, size=10)
            ws_imp.cell(current_row_imp,3).value = "RAZÓN SOCIAL:"
            cell_style(ws_imp.cell(current_row_imp,3), C_HDR_BG, True, C_WHITE, size=10)
            ws_imp.cell(current_row_imp,4).value = ", ".join(rns)
            cell_style(ws_imp.cell(current_row_imp,4), C_HDR_BG, False, C_WHITE, size=10)
            ws_imp.cell(current_row_imp,5).value = "EJERCICIO:"
            cell_style(ws_imp.cell(current_row_imp,5), C_HDR_BG, True, C_WHITE, size=10)
            ws_imp.cell(current_row_imp,6).value = ej
            cell_style(ws_imp.cell(current_row_imp,6), C_HDR_BG, False, C_WHITE, size=10)

            groups_period = defaultdict(list)
            for r in rfc_ej_records:
                pe = r.get("periodo", "")
                groups_period[pe].append(r)
                
            vigentes = []
            anteriores = []
            for pe, recs_pe in groups_period.items():
                sorted_recs = sorted(recs_pe, key=get_date_key, reverse=True)
                if sorted_recs:
                    vigentes.append(sorted_recs[0])
                    anteriores.extend(sorted_recs[1:])
                
            vigentes.sort(key=get_period_sort_key)
            anteriores.sort(key=get_period_sort_key)
            
            ws_imp.cell(current_row_imp + 2, 1, value="RELACIÓN DE PAGOS VIGENTES").font = Font(bold=True, size=11, color="37474F")
            t1_end = render_tax_table(ws_imp, current_row_imp + 3, imp, vigentes, imp)
            
            if anteriores:
                ws_imp.cell(t1_end + 3, 1, value="RELACIÓN DE PAGOS ANTERIORES (DEDUPLICADOS/REEMPLAZADOS)").font = Font(bold=True, size=11, color="37474F")
                t1_end = render_tax_table(ws_imp, t1_end + 4, imp, anteriores, imp)
                
            current_row_imp = t1_end + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Leer Excel para continuar ─────────────────────────────────────────────────
def read_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Pagos de Impuestos" in wb.sheetnames:
        ws = wb["Pagos de Impuestos"]
    else:
        ws = wb.active
    records = []

    # Detecta encabezados en fila 3
    imp_col_map = {}   # col_index -> (imp, sub_key)
    hdr_col_map = {}   # col_index -> hdr_key
    tail_col_map = {}  # col_index -> tail_key

    row3 = [c.value for c in ws[3]]
    row2 = [c.value for c in ws[2]]

    current_imp = None
    for ci, (h2, h3) in enumerate(zip(row2, row3)):
        if h2 and h2 in ALL_IMPUESTOS:
            current_imp = h2
        # Si h3 es None pero h2 tiene valor (celda combinada verticalmente), usar h2
        effective_h = h3 if h3 is not None else h2
        if effective_h in HDR_COLS:
            hdr_col_map[ci] = HDR_KEYS[HDR_COLS.index(effective_h)]
        elif effective_h in SUB_COLS and current_imp:
            imp_col_map[ci] = (current_imp, SUB_KEYS[SUB_COLS.index(effective_h)])
        elif effective_h in TAIL_COLS:
            tail_col_map[ci] = TAIL_KEYS[TAIL_COLS.index(effective_h)]

    for row in ws.iter_rows(min_row=4, values_only=True):
        if all(v is None for v in row): continue
        if isinstance(row[0], str) and "TOTAL" in str(row[0]).upper(): continue
        rec = {}
        for ci, v in enumerate(row):
            if ci in hdr_col_map:
                rec[hdr_col_map[ci]] = str(v) if v is not None else ""
            elif ci in imp_col_map:
                imp, sk = imp_col_map[ci]
                rec[f"{imp}|{sk}"] = float(v) if isinstance(v,(int,float)) else v
            elif ci in tail_col_map:
                tk = tail_col_map[ci]
                rec[tk] = float(v) if isinstance(v,(int,float)) and tk in ("importe_total",) else (str(v) if v else "")
        # garantiza que todos los campos de impuesto existan
        for imp in ALL_IMPUESTOS:
            for sk in SUB_KEYS:
                rec.setdefault(f"{imp}|{sk}", None)
        records.append(rec)

    return records


# ── Deduplicación — clave: num_operacion_declaracion ──────────────────────────
def tiene_banco(rec):
    """True si el comprobante ya tiene fecha de pago bancario."""
    return bool(rec.get("fecha_pago"))

def merge_records(existing_list, incoming_list):
    """
    Combina dos listas de registros aplicando deduplicación por num_operacion_declaracion.
    Reglas:
      - Hay duplicado + entrante tiene banco  → reemplaza al existente
      - Hay duplicado + entrante NO tiene banco → descarta el entrante
      - Sin duplicado → agrega normalmente
    """
    result = list(existing_list)  # copia

    for new_rec in incoming_list:
        key = new_rec.get("num_operacion_declaracion", "").strip()
        if not key:
            # Sin clave única, simplemente agrega (no podemos deduplicar)
            result.append(new_rec)
            continue

        # Buscar si ya existe en result
        idx_found = None
        for i, existing in enumerate(result):
            if existing.get("num_operacion_declaracion", "").strip() == key:
                idx_found = i
                break

        if idx_found is None:
            # No existe duplicado → agregar
            result.append(new_rec)
        else:
            existing = result[idx_found]
            new_has_banco   = tiene_banco(new_rec)
            exist_has_banco = tiene_banco(existing)

            if new_has_banco and not exist_has_banco:
                # El nuevo tiene banco, el existente no → reemplazar
                result[idx_found] = new_rec
            # En cualquier otro caso se descarta el nuevo (ya tenemos uno igual o mejor)

    return result


# ── Session store ─────────────────────────────────────────────────────────────
SESSION_DATA: dict = {}


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/session/new", methods=["POST"])
def new_session():
    sid = str(uuid.uuid4())
    SESSION_DATA[sid] = []
    return jsonify({"session_id": sid})

@app.route("/api/session/<sid>")
def get_session(sid):
    recs = SESSION_DATA.get(sid, [])
    return jsonify({"session_id": sid, "records": recs, "count": len(recs)})

@app.route("/api/upload-pdf", methods=["POST"])
def upload_pdf():
    sid = request.form.get("session_id")
    if not sid: return jsonify({"error": "Sesión inválida"}), 400
    SESSION_DATA.setdefault(sid, [])
    files = request.files.getlist("files")
    if not files: return jsonify({"error": "Sin archivos"}), 400
    
    import json
    paths_str = request.form.get("file_paths", "[]")
    try:
        file_paths = json.loads(paths_str)
    except:
        file_paths = []

    print(f"\\n--- UPLOAD PDF ---")
    print(f"Archivos recibidos: {len(files)}")
    new_recs, errors = [], []
    for i, f in enumerate(files):
        if not f.filename.lower().endswith(".pdf"):
            errors.append(f"{f.filename}: solo PDF"); continue
        try:
            import os
            basename = os.path.basename(f.filename)
            # Usar la ruta capturada del navegador si está disponible
            route = file_paths[i] if i < len(file_paths) else f.filename
            rec = parse_pdf(f.read(), basename, route)
            new_recs.append(rec)
            print(f"Parsed {f.filename}: Num={rec.get('num_operacion_declaracion')}, Banco={tiene_banco(rec)}")
        except Exception as e:
            errors.append(f"{f.filename}: {e}")

    print(f"Total parsed correctly: {len(new_recs)}")
    
    # Deduplicar: primero entre los propios nuevos, luego contra la sesión
    new_recs = merge_records([], new_recs)          # dedup entre lote entrante
    print(f"After self-dedup: {len(new_recs)}")
    
    original_count = len(SESSION_DATA[sid])
    SESSION_DATA[sid] = merge_records(SESSION_DATA[sid], new_recs)
    added = len(SESSION_DATA[sid]) - original_count
    
    print(f"After session dedup: {len(SESSION_DATA[sid])} (Added: {added})")

    return jsonify({"new_count": added, "total_count": len(SESSION_DATA[sid]),
                    "new_records": new_recs, "errors": errors})

@app.route("/api/upload-excel", methods=["POST"])
def upload_excel():
    sid = request.form.get("session_id")
    if not sid: return jsonify({"error": "Sesión inválida"}), 400
    SESSION_DATA.setdefault(sid, [])
    f = request.files.get("file")
    if not f: return jsonify({"error": "Sin archivo"}), 400
    try:
        recs = read_excel(f.read())
        original_count = len(SESSION_DATA[sid])
        SESSION_DATA[sid] = merge_records(SESSION_DATA[sid], recs)
        added = len(SESSION_DATA[sid]) - original_count
        return jsonify({"loaded_count": added, "total_count": len(SESSION_DATA[sid]), "records": SESSION_DATA[sid]})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/delete-record", methods=["POST"])
def delete_record():
    d = request.json
    sid, idx = d.get("session_id"), d.get("index")
    SESSION_DATA.setdefault(sid, [])
    recs = SESSION_DATA[sid]
    if 0 <= idx < len(recs): recs.pop(idx)
    return jsonify({"total_count": len(recs), "records": recs})

@app.route("/api/export-excel", methods=["POST"])
def export_excel():
    d = request.json
    sid = d.get("session_id")
    if not sid or sid not in SESSION_DATA:
        return jsonify({"error": "Sesión inválida"}), 400
    recs = SESSION_DATA[sid]
    if not recs: return jsonify({"error": "Sin registros"}), 400
    buf = io.BytesIO(generate_excel(recs))
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"pagos_impuestos_{ts}.xlsx")

@app.route("/api/clear-session", methods=["POST"])
def clear_session():
    sid = request.json.get("session_id")
    if sid in SESSION_DATA: SESSION_DATA[sid] = []
    return jsonify({"ok": True})

if __name__ == "__main__":
    app.run(debug=True, port=5050, use_reloader=False)
